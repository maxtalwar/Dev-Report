import openpyxl, requests, re, time, flask_wrapper, tools, signal, excelFileStorage, os, pandas
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from bs4 import BeautifulSoup
from urllib.parse import urlparse
from flask import Flask, render_template
from contextlib import contextmanager
from threading import active_count
from colorama import Fore, Style

# Create a session object
session = requests.Session()
session.max_redirects = 10

# Create an empty cache dictionary
cache = {}
url_cache = tools.load_cache_from_json("url_to_github_cache.json")

# Open the Excel file
wb = openpyxl.load_workbook('max_tasks.xlsx')

# Select the sheet that contains the data
sheet = wb['Task 4 - Empty ecosystems']

# Get the total number of rows in the sheet
total_rows = 1939
rows_scanned = 0

# Create an empty list to store the GitHub URLs
github_urls = []

class TimeoutException(Exception): pass

@contextmanager
def time_limit(seconds):
    def signal_handler(signum, frame):
        raise TimeoutException("Timed out!")
    signal.signal(signal.SIGALRM, signal_handler)
    signal.alarm(seconds)
    try:
        yield
    finally:
        signal.alarm(0)

def get_html(website_url):
    # Check if the website URL is in the cache
    if website_url in cache:
        # If it is, get the HTML code from the cache
        html = cache[website_url]
    else:
    # Try to send a request to the website to get the HTML code
        try:
            response = session.get(website_url)
            if response.status_code != 200:
                return ''
            html = response.text
            # Store the HTML code in the cache
            cache[website_url] = html
        except KeyboardInterrupt:
            return ''
        # If a connection cannot be established, skip the iteration
        except requests.exceptions.ConnectionError:
            return ''
        # If the website gives too many redirects, skip the iteration
        except requests.exceptions.TooManyRedirects:
            return ''
        except TimeoutException:
            return ''
        except:
            pass
    
    return html

def find_github_urls_beautifulsoup(html):
    # Create an empty list to store the GitHub URLs
    github_urls = []

    try:
        with time_limit(240):
            # Use BeautifulSoup to parse the HTML code
            soup = BeautifulSoup(html, 'html.parser')

            # Find all links in the HTML code
            links = soup.find_all('a')

            # Iterate through each link
            for link in links:
                # Get the link's href attribute
                href = link.get('href')
                parsed_url = urlparse(href)

                # Filter out useless URLs
                if not href or not parsed_url.scheme or not parsed_url.netloc or parsed_url.scheme not in ["http", "https"]:
                    continue

                # If the link points to a GitHub URL, add it to the list
                if href and 'github.com' in href:
                    github_urls.append(href)
                # If the link does not point to a GitHub URL, follow it and check if it resolves to a GitHub URL
                else:
                    try:
                        final_url = session.get(href).url
                    except requests.exceptions.ConnectionError:
                        continue

                    if "github.com" in final_url:
                        github_urls.append(final_url)
    except TimeoutException as e:
        pass
    except KeyboardInterrupt:
        pass
    except requests.exceptions.ContentDecodingError:
        pass
    except requests.exceptions.TooManyRedirects:
        pass
    except requests.exceptions.InvalidURL:
        pass
    except:
        pass

    return github_urls

def find_github_urls_regex(html):
    # Use a regular expression to find all GitHub URLs in the HTML code
    github_urls = re.findall(r'https://github\.com/[\w-]+/$', html)

    return github_urls

def get_github_type(url):
    if re.match(r'https://github\.com/[\w-]+/[\w-]+', url):
        url_type = 'repository'
    elif re.match(r'https://github\.com/[\w-]+', url):
        url_type = 'user'
    else:
        url_type = 'organization'
    
    return url_type

def choose_url(urls):
    url_types = {
        'organization': 3,
        'repository': 2,
        'user': 1
    }

    priority = 0
    chosen_url = None

    for url in urls:
        url_type = get_github_type(url)
        if url_types[url_type] > priority:
            # If it does, update the priority and the chosen URL
            priority = url_types[url_type]
            chosen_url = url

    return chosen_url

def display_info(rows_scanned, row_num, website_url, current_function):
    # print(f'[{now} Row {row_num} {len(discovered_urls)}] Progress: {progress:.2f}%')

    # Clear the page
    os.system('clear')

    # Calculate the progress percentage
    progress = rows_scanned / total_rows * 100

    # Get the current date and time
    now = pandas.to_datetime(datetime.now()).round('1s')

    # Print the progress percentage, row number, and date and time, and URL
    print(f'Progress: ' + Fore.GREEN + str(round(progress, 2)) + "%" + Style.RESET_ALL)
    print(f'Start Time: {now}')
    print(f'Row: {row_num}')
    print(f'Website: {website_url}')
    print(f'Function: {current_function}')

if __name__ == "__main__":
    # Create a thread pool with 4 worker threads
    with ThreadPoolExecutor(max_workers=4) as executor:
        # Iterate through each row in the sheet
        for row_num, row in enumerate(sheet.iter_rows(min_row=4, max_col=5, max_row=sheet.max_row), start=1):
            # Get the website URL from column E (column 5)
            website_url = row[4].value

            # Skip the iteration if the website URL is None or an empty string
            if not website_url:
                continue

            # update info on console
            display_info(rows_scanned, row_num, website_url, current_function="reference_cache")

            # check to see if website already has a stored github link, and if so just add the one that has already been stored
            if website_url in url_cache.keys():
                github_url = url_cache[website_url]
                discovered_urls = []
            else:
                try:
                    # get HTML
                    display_info(rows_scanned, row_num, website_url, current_function="get_html")
                    future = executor.submit(get_html, website_url)
                    html = future.result()
                except KeyboardInterrupt:
                    html = ""

                # find all github links
                display_info(rows_scanned, row_num, website_url, current_function="find_github_urls")
                discovered_urls = find_github_urls_regex(html)

                if len(discovered_urls) == 0:
                    discovered_urls = find_github_urls_beautifulsoup(html)

                # choose the URL out of the given URLs
                display_info(rows_scanned, row_num, website_url, current_function="choose_url")
                github_url = choose_url(discovered_urls)

                url_cache[website_url] = github_url

            github_urls.append(github_url)

            tools.save_cache_to_json(url_cache, "url_to_github_cache.json")

            rows_scanned += 1

            # Add a delay of 0.05 seconds between iterations to avoid sending too many requests in a short period of time
            time.sleep(0.05)

excelFileStorage.update_excel_file()