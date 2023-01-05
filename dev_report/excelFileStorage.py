import openpyxl
import json

def update_excel_file():
    # Load the JSON file
    with open("url_to_github_cache.json", "r") as f:
        url_to_github_cache = json.load(f)

    # Open the Excel file
    wb = openpyxl.load_workbook("max_tasks.xlsx")

    # Select the sheet that contains the data
    sheet = wb["Task 4 - Empty ecosystems"]

    # Iterate through each row in the sheet
    for row in sheet.iter_rows(min_row=4, min_col=5, max_col=5):
        # Get the value in column E (website URL)
        website_url = str(row[0].value)

        # Check if the website URL is in the url_to_github_cache dictionary
        if website_url in url_to_github_cache:
            # If it is, get the corresponding GitHub URL
            github_url = str(url_to_github_cache[website_url])
            if github_url != "None":
                # Update the cell in column F with the GitHub URL
                sheet.cell(row=row[0].row, column=6).value = github_url

    # Save the updated Excel file
    wb.save("max_tasks.xlsx")

if __name__ == "__main__":
    update_excel_file()